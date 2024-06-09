import flask,json
import re 
from openpyxl import load_workbook
import pandas as pd
from flask import render_template_string
import datetime 
import os
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address



def generate_html_page(data, file_path = '展示html/招聘信息展示.html'):
    html_table = """
    <table border="1">
        <tr>
            <th>公司/单位名称</th>
            <th>校招类型</th>
            <th>官方标题</th>
            <th>公告原文链接</th>
            <th>网申/投递地址</th>
            <th>开启时间</th>
            <th>截止时间</th>
            <th>地点</th>
        </tr>
    """
    for item in data:
        html_table += f"""
        <tr>
            <td>{item['公司/单位名称']}</td>
            <td>{item['校招类型']}</td>
            <td>{item['官方标题']}</td>
            <td><a href="{item['公告原文链接']}">点击查看</a></td>
            <td>{item['网申/投递地址']}</td>
            <td>{item['开启时间']}</td>
            <td>{item['截止时间']}</td>
            <td>{item['地点']}</td>
        </tr>
        """

    html_table += "</table>"

    html_template = f"""
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
        <title>招聘信息展示</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                background-color: #f8f9fa;
            }}
            h1 {{
                text-align: center;
                color: #333;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
                background-color: #fff;
                box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            }}
            th, td {{
                padding: 10px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }}
            th {{
                background-color: #f2f2f2;
                font-weight: bold;
            }}
            tr:hover {{
                background-color: #f1f1f1;
            }}
            a {{
                color: #007bff;
                text-decoration: none;
            }}
            a:hover {{
                text-decoration: underline;
            }}
        </style>
    </head>
    <body>
        <h1>招聘信息</h1>
        {html_table}
    </body>
    </html>
    """

    # 将生成的 HTML 保存到文件
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(html_template)




server = flask.Flask(__name__)
# 配置限流器
limiter = Limiter(
    get_remote_address,
    app=server,
    default_limits=["200 per day", "50 per hour"]  # 设置默认限流策略
)


@server.route('/index',methods=['post'])
def index():
    # # 加载工作簿
    # workbook = load_workbook(filename='秋招信息表.xlsx')

    # # 选择工作表
    # sheet = workbook['工作表1']

    # # 读取整个工作表数据
    # resp = ""
    # for row in sheet.iter_rows(min_row=2, max_row=50, max_col=9, values_only=True):
    #     print(row)
    #     resp += str(row)

    # 获取查询参数
    data = flask.request.get_json()
    company_name = data.get('company_name', '')
    job_type = data.get('job_type', '')
    open_time = data.get('open_time', '')
    fact_types = data.get('fact_types','')
    locations = data.get('locations','')
    # print("company_name"+company_name)
    # print("job_type"+job_type)
    # print("open_time"+open_time)
    # 加载工作簿
    workbook = load_workbook(filename='秋招信息表.xlsx')

    # 选择工作表
    sheet = workbook['工作表1']

    # 读取整个工作表数据并进行模糊搜索
    resp = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = ' '.join(map(str, row))  # 将行数据拼接成字符串
        # if re.search(query, row_data, re.IGNORECASE):  
        # 使用正则表达式进行模糊搜索
        if (not company_name or re.search(company_name, str(row[1]), re.IGNORECASE)) and \
           (not job_type or re.search(job_type, str(row[2]), re.IGNORECASE)) and \
           (not fact_types or re.search(fact_types, str(row[9]), re.IGNORECASE)) and \
           (not locations or re.search(locations, str(row[8]), re.IGNORECASE)) and \
           (not open_time or re.search(open_time, str(row[6]), re.IGNORECASE)):
            print(row)
            resp.append({
                # '收录时间': row[0],
                # '收录时间': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime.datetime) else row[0],
                '公司/单位名称': row[1],
                '校招类型': row[2],
                '官方标题': row[3],
                '公告原文链接': row[4],
                '网申/投递地址': row[5],
                '开启时间': row[6].strftime('%Y-%m-%d') if isinstance(row[6], datetime.datetime) else row[6],
                '截止时间': row[7].strftime('%Y-%m-%d') if isinstance(row[7], datetime.datetime) else row[7],
                '地点': row[8],
                '行业分类': row[9] if len(row) > 9 else '',
                '企业性质': row[10] if len(row) > 10 else '',
                # '备注': row[11] if len(row) > 11 else ''
            })
    
    res={'msg':str(resp),'msg_dode':0}
    generate_html_page(resp)
    # return json.dumps(res, ensure_ascii=False)
    return res

@server.route('/biaoge',methods=['get'])
def biaoge():
    try:
        # 读取 Excel 文件
        df = pd.read_excel('秋招信息表.xlsx')

        # 将 DataFrame 转换为 HTML 表格
        html_table = df.to_html(classes='table table-striped', index=False)

        # 使用简单的 HTML 模板展示表格
        html_template = f"""
        <!doctype html>
        <html lang="en">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
            <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.3.1/css/bootstrap.min.css">
            <title>Excel Viewer</title>
        </head>
        <body>
            <div class="container">
            <h1 class="mt-5">Excel File Viewer</h1>
            {html_table}
            </div>
        </body>
        </html>
        """
        return render_template_string(html_template)
    except FileNotFoundError:
        return jsonify({"error": "文件未找到"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@server.route('/recently',methods=['post'])
def recently():
    data = flask.request.get_json()
    open_time = data.get('open_time', '')
    input_date = datetime.datetime.strptime(open_time, '%Y-%m-%d')
    time_num = data.get('time_num', '')
    start_date = input_date - datetime.timedelta(days=time_num)
    end_date = input_date

    workbook = load_workbook(filename='秋招信息表.xlsx')
    sheet = workbook['工作表1']
    resp = []

    job_type = "2025届校招/秋招"

    for row in sheet.iter_rows(min_row=3, values_only=True):
        # print(row[6])
        print("-------------------------------------------------------")
        if str(row[6]) != "None":
            open_date = datetime.datetime.strptime(str(row[6]), '%Y-%m-%d %H:%M:%S')
        else :
            open_date = datetime.datetime.strptime("1999-01-01 00:00:00", '%Y-%m-%d %H:%M:%S')
        if (not job_type or re.search(job_type, str(row[2]), re.IGNORECASE)):
            if start_date <= open_date <= end_date:
                print("open_date"+str(open_date))
                # print("start_date"+str(start_date))
                # print("end_date"+str(end_date))
                resp.append({
                    '公司/单位名称': row[1],
                    '校招类型': row[2],
                    '官方标题': row[3],
                    '公告原文链接': row[4],
                    '网申/投递地址': row[5],
                    '开启时间': row[6].strftime('%Y-%m-%d') if isinstance(row[6], datetime.datetime) else row[6],
                    '截止时间': row[7].strftime('%Y-%m-%d') if isinstance(row[7], datetime.datetime) else row[7],
                    '地点': row[8],
                    '行业分类': row[9] if len(row) > 9 else '',
                    '企业性质': row[10] if len(row) > 10 else '',
                })
    generate_html_page(resp)
    res = {'msg': resp, 'msg_code': 0}
    return res


@server.route('/show', methods=['get'])
@limiter.limit("10 per minute")  # 为这个特定的路由设置限流策略
def show():
    file_path = '展示html/招聘信息展示.html'
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            html_content = file.read()
        return render_template_string(html_content)
    else:
        return "没有可显示的数据", 404


if __name__ == '__main__':
    server.run(host='0.0.0.0',port=5000,debug=True)
